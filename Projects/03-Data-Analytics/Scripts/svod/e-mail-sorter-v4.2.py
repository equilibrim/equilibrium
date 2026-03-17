import os
import re
import shutil
from datetime import datetime
import openpyxl
import logging
from pathlib import Path
import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import sys

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

class ReportSorter:
    def __init__(self, source_folder, output_folder, report_names_file, interactive=False):
        self.source_folder = source_folder
        self.output_folder = output_folder
        self.report_names_file = report_names_file
        self.interactive = interactive
        os.makedirs(output_folder, exist_ok=True)
        # Основные форматы
        self.supported_formats = ['.xlsx', '.xls', '.pdf', '.docx', '.doc']
        # Словари для хранения: {ключ_поиска: (название_папки, тип_поиска)}
        # тип_поиска: 'content' или 'filename'
        self.search_to_folder = {}
        self.found_folders = set()
        # Статистика
        self.stats = {
            'total_files': 0,
            'processed': 0,
            'sorted': 0,
            'not_found': 0,
            'errors': 0,
            'moved': 0,
            'interactive_choices': 0,
            'exact_matches': 0,
            'name_matches': 0,
            'new_keys_added': 0
        }
        # Для хранения неотсортированных файлов
        self.unsorted_files = []
        self.all_files_original = []
        # Лог файл
        self.log_file = os.path.join(self.output_folder, "детальный_лог.txt")
        with open(self.log_file, 'w', encoding='utf-8') as f:
            f.write(f"Лог сортировки - {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}\n")
            f.write("="*60 + "\n")

    def log_detail(self, message):
        """Запись детального лога"""
        with open(self.log_file, 'a', encoding='utf-8') as f:
            f.write(f"{datetime.now().strftime('%H:%M:%S')} - {message}\n")

    def extract_organization_from_path(self, file_path, rel_path):
        """Извлечение названия организации из пути к файлу"""
        try:
            parts = rel_path.split(os.sep)
            if len(parts) >= 2:
                org_name = parts[0]
                org_name = re.sub(r'[<>:"/\\|?*]', '_', org_name)
                org_name = org_name.strip('_')
                if len(org_name) > 30:
                    org_name = org_name[:20] + "..." + org_name[-5:]
                return org_name if org_name else "Неизвестно"
            filename = os.path.basename(file_path)
            patterns = [
                r'от\s+([^_\-.]+)',
                r'([А-Я][а-я]+)\s+отчет',
                r'([А-Я]+[\w\s]+)_отчет',
            ]
            for pattern in patterns:
                match = re.search(pattern, filename, re.IGNORECASE)
                if match:
                    org_name = match.group(1).strip()
                    org_name = re.sub(r'[<>:"/\\|?*]', '_', org_name)
                    return org_name[:30]
            return "Неизвестно"
        except Exception:
            return "Неизвестно"

    def load_report_names(self):
        """Загрузка ключей поиска и названий папок из файла"""
        print(f"\n📋 Загрузка настроек из: {self.report_names_file}")
        if not os.path.exists(self.report_names_file):
            print(f"❌ Файл не найден: {self.report_names_file}")
            return False

        try:
            with open(self.report_names_file, 'r', encoding='utf-8') as f:
                lines = [line.strip() for line in f if line.strip()]

            print(f"📄 Загружено строк: {len(lines)}")
            for line in lines:
                parts = line.split('|', 2)
                if len(parts) == 3:
                    search_key = parts[0].strip()
                    folder_name = parts[1].strip()
                    search_type = parts[2].strip().lower()
                    if search_key and folder_name and search_type in ['content', 'filename']:
                        self.search_to_folder[search_key] = (folder_name, search_type)
                elif len(parts) == 2:
                    search_key = parts[0].strip()
                    folder_name = parts[1].strip()
                    if search_key and folder_name:
                        self.search_to_folder[search_key] = (folder_name, 'content')
                else:
                    search_key = line.strip()
                    if search_key:
                        self.search_to_folder[search_key] = (search_key, 'content')

            print(f"✅ Загружено ключей поиска: {len(self.search_to_folder)}")
            print(f"✅ Будут созданы папки: {len(set([v[0] for v in self.search_to_folder.values()]))}")
            debug_file = os.path.join(self.output_folder, "настройки_поиска.txt")
            with open(debug_file, 'w', encoding='utf-8') as f:
                f.write("НАСТРОЙКИ ПОИСКА И СОРТИРОВКИ:\n")
                f.write("="*80 + "\n")
                f.write("Формат: 'КЛЮЧ_ПОИСКА | НАЗВАНИЕ_ПАПКИ | ТИП_ПОИСКА'\n")
                f.write("ТИП_ПОИСКА: 'content' или 'filename'\n")
                f.write("ИЛИ просто 'КЛЮЧ_ПОИСКА' (ключ = имя папки, поиск в содержимом)\n")
                f.write("="*80 + "\n")
                f.write("📋 СПИСОК КЛЮЧЕЙ ДЛЯ ПОИСКА:\n")
                for search_key, (folder_name, search_type) in sorted(self.search_to_folder.items()):
                    f.write(f"\n🔍 Ищем: '{search_key}' (тип: {search_type})")
                    if search_key != folder_name:
                        f.write(f" → 📁 Папка: '{folder_name}'")
                    f.write("\n")
            return True
        except Exception as e:
            print(f"❌ Ошибка загрузки: {e}")
            return False

    def save_report_names(self):
        """Сохранение ключей поиска в файл"""
        try:
            with open(self.report_names_file, 'w', encoding='utf-8') as f:
                for search_key, (folder_name, search_type) in self.search_to_folder.items():
                    if search_key == folder_name and search_type == 'content':
                        f.write(f"{search_key}\n")
                    else:
                        f.write(f"{search_key} | {folder_name} | {search_type}\n")
            print(f"✅ Настройки сохранены в файл: {self.report_names_file}")
            return True
        except Exception as e:
            print(f"❌ Ошибка сохранения настроек: {e}")
            return False

    def search_exact_in_excel(self, file_path, filename):
        """ТОЧНЫЙ поиск ключей в содержимом Excel файла"""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            all_text_lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=1, max_row=500, min_col=1, max_col=20, values_only=True):
                    row_texts = [str(cell).strip() for cell in row if cell]
                    if row_texts:
                        row_line = ' '.join(row_texts)
                        all_text_lines.append(row_line)
            wb.close()

            if not all_text_lines:
                return None

            for search_key, (folder_name, search_type) in self.search_to_folder.items():
                if search_type == 'content':
                    for line in all_text_lines:
                        if search_key in line:
                            return folder_name
            return None
        except Exception as e:
            self.log_detail(f"Ошибка чтения Excel {filename}: {e}")
            return None

    def search_exact_in_pdf(self, file_path, filename):
        """ТОЧНЫЙ поиск ключей в содержимом PDF"""
        try:
            import PyPDF2
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                pdf_lines = []
                for page in pdf_reader.pages:
                    text = page.extract_text()
                    if text:
                        lines = text.split('\n')
                        for line in lines:
                            line_clean = line.strip()
                            if line_clean:
                                pdf_lines.append(line_clean)

                if pdf_lines:
                    for search_key, (folder_name, search_type) in self.search_to_folder.items():
                        if search_type == 'content':
                            for line in pdf_lines:
                                if search_key in line:
                                    return folder_name
            return None
        except Exception as e:
            self.log_detail(f"Ошибка PDF {filename}: {e}")
            return None

    def search_in_filename(self, filename):
        """Поиск ключей в имени файла, учитывая тип поиска"""
        name_without_ext = os.path.splitext(filename)[0]
        clean_name = re.sub(r'[_\-.]', ' ', name_without_ext.lower())

        for search_key, (folder_name, search_type) in self.search_to_folder.items():
            if search_type == 'filename':
                if search_key.lower() in clean_name:
                    return folder_name
        return None

    def identify_report_type(self, file_path):
        """Поиск ТОЛЬКО в содержимом файлов (оригинальная логика)"""
        filename = os.path.basename(file_path)
        file_ext = os.path.splitext(filename)[1].lower()

        if file_ext in ['.xlsx', '.xls']:
            return self.search_exact_in_excel(file_path, filename)
        elif file_ext == '.pdf':
            return self.search_exact_in_pdf(file_path, filename)
        elif file_ext in ['.docx', '.doc']:
            return None
        return None

    def identify_report_type_with_filename(self, file_path):
        """Поиск в содержимом файлов И в именах файлов (для ресортировки)"""
        filename = os.path.basename(file_path)
        file_ext = os.path.splitext(filename)[1].lower()

        folder_name = self.search_in_filename(filename)
        if folder_name:
            self.stats['name_matches'] += 1
            return folder_name

        if file_ext in ['.xlsx', '.xls']:
            return self.search_exact_in_excel(file_path, filename)
        elif file_ext == '.pdf':
            return self.search_exact_in_pdf(file_path, filename)
        elif file_ext in ['.docx', '.doc']:
            return None
        return None

    def get_interactive_choice(self, filename, file_ext, file_path, organization):
        """Интерактивный выбор для нераспознанных файлов"""
        print(f"\n{'='*60}")
        print(f"❓ ФАЙЛ НЕ РАСПОЗНАН: {filename}")
        print(f"   Полный путь: {file_path}")
        print(f"   Организация: {organization}")
        print(f"   Формат: {file_ext}")
        print(f"{'-'*60}")
        existing_folders = sorted(list(self.found_folders))
        if existing_folders:
            print("Существующие папки:")
            for i, folder in enumerate(existing_folders[:20], 1):
                print(f"  {i:2}. {folder}")
            if len(existing_folders) > 20:
                print(f"  ... и еще {len(existing_folders) - 20} папок")

        print("\nВыберите действие:")
        print("  1. Создать новую папку")
        if existing_folders:
            print("  2. Выбрать из существующих папок")
        print("  3. Поместить в 'НЕ_СОРТИРОВАННЫЕ'")
        print("  4. Перейти в режим сканирования по имени файла") # Новое сообщение
        print("  5. Добавить новый ключ поиска в содержимое")
        print("  6. Добавить новый ключ поиска по имени файла")
        print("  7. Просмотреть содержимое файла")

        while True:
            choice = input("\nВаш выбор: ").strip()
            if choice == '1':
                folder_name = input("Введите название новой папки: ").strip()
                if folder_name:
                    return folder_name
                else:
                    print("Название папки не может быть пустым!")
            elif choice == '2' and existing_folders:
                try:
                    folder_num = int(input(f"Введите номер папки (1-{len(existing_folders)}): "))
                    if 1 <= folder_num <= len(existing_folders):
                        return existing_folders[folder_num - 1]
                    else:
                        print(f"Неверный номер! Введите от 1 до {len(existing_folders)}")
                except ValueError:
                    print("Введите число!")
            elif choice == '3':
                return "НЕ_СОРТИРОВАННЫЕ"
            elif choice == '4': # Новая логика для выбора 4
                print("\n🔄 Переход в режим сканирования по имени файла...")
                # Выполняем сортировку по имени файла для всех неотсортированных
                self.scan_and_sort_by_filename()
                print("🔄 Завершена сортировка по имени файла.")
                # Возвращаемся к выбору для текущего файла
                return self.get_interactive_choice(filename, file_ext, file_path, organization)
            elif choice == '5':
                result = self.add_new_search_key(file_path, filename, file_ext, search_type='content')
                if result:
                    return result
                else:
                    print("Продолжаем выбор папки для текущего файла...")
                    continue
            elif choice == '6':
                result = self.add_new_search_key(file_path, filename, file_ext, search_type='filename')
                if result:
                    return result
                else:
                    print("Продолжаем выбор папки для текущего файла...")
                    continue
            elif choice == '7':
                self.preview_file_content(file_path, file_ext)
                continue
            else:
                print("Неверный выбор! Введите 1, 2, 3, 4, 5, 6 или 7")


    def add_new_search_key(self, file_path, filename, file_ext, search_type='content'):
        """Добавление нового ключа поиска с автоматической ресортировкой"""
        print(f"\n➕ ДОБАВЛЕНИЕ НОВОГО КЛЮЧА ПОИСКА ({'в содержимом' if search_type == 'content' else 'в имени файла'})")
        print(f"Файл: {filename}")

        print("\nСодержимое файла (первые 200 символов):")
        content_preview = self.get_file_preview(file_path, file_ext, max_chars=200)
        print(f"  {content_preview}")

        if search_type == 'content':
            print("\nВы можете:")
            print("  1. Ввести текст вручную")
            print("  2. Выбрать текст из содержимого файла")
            choice = input("Ваш выбор (1 или 2): ").strip()
            search_key = ""
            if choice == '2':
                print("\nСодержимое файла для выбора текста:")
                full_preview = self.get_file_preview(file_path, file_ext, max_chars=1000)
                lines = full_preview.split('\n')
                print("="*60)
                for i, line in enumerate(lines[:20], 1):
                    print(f"{i:2}. {line}")
                print("="*60)
                try:
                    line_num = int(input(f"\nВыберите номер строки (1-{min(20, len(lines))}): "))
                    if 1 <= line_num <= len(lines):
                        selected_line = lines[line_num-1]
                        print(f"\nВыбранная строка: '{selected_line}'")
                        search_key = input("  Ключ поиска: ").strip()
                except (ValueError, IndexError):
                    print("Неверный выбор, вводите текст вручную.")
                    search_key = input("\nВведите текст для поиска в файлах: ").strip()
            else:
                search_key = input("\nВведите текст для поиска в файлах: ").strip()
        else: # search_type == 'filename'
            print("\nВведите текст, который должен содержаться в имени файла:")
            print(f"  Текущее имя: {filename}")
            search_key = input("  Ключ поиска в имени файла: ").strip()

        if not search_key:
            print("Ключ поиска не может быть пустым!")
            return None

        print("\nВведите название папки для этого ключа:")
        print("  1. Создать новую папку")
        print("  2. Выбрать из существующих")
        folder_choice_input = input("  Ваш выбор (1 или 2): ").strip()
        folder_name = ""

        existing_folders = sorted(list(self.found_folders))
        if folder_choice_input == '2' and existing_folders:
            print("\nСуществующие папки:")
            for i, folder in enumerate(existing_folders, 1):
                print(f"  {i:2}. {folder}")
            try:
                folder_num = int(input(f"\nВведите номер папки (1-{len(existing_folders)}): "))
                if 1 <= folder_num <= len(existing_folders):
                    folder_name = existing_folders[folder_num - 1]
                else:
                    print(f"Неверный номер! Введите от 1 до {len(existing_folders)}")
                    return None
            except ValueError:
                print("Введите число!")
                return None
        elif folder_choice_input == '1':
            folder_name = input("\nВведите название новой папки: ").strip()
        else:
            print("Неверный выбор!")
            return None

        if not folder_name:
            print("Название папки не может быть пустым!")
            return None

        self.search_to_folder[search_key] = (folder_name, search_type)
        self.stats['new_keys_added'] += 1
        print(f"\n✅ Добавлен ключ поиска: '{search_key}' → папка '{folder_name}' (тип поиска: {search_type})")

        self.save_report_names()

        # Поиск в текущем файле
        print(f"\n🔍 Поиск нового ключа в текущем файле...")
        found_in_current = self.find_folder_by_newest_key(file_path, search_type)
        if found_in_current:
            print(f"✅ Найден ключ в текущем файле! Будет перемещён в: '{found_in_current}'")
        else:
            print("⚠️  Ключ не найден в текущем файле.")

        # === ВСЕГДА выполняем автоматическую ресортировку ===
        if self.unsorted_files:
            print(f"\n🔄 Запуск автоматической ресортировки неотсортированных файлов...")
            sorted_count = self.rescan_unsorted_by_search_type(search_key, search_type)
            print(f"✅ Ресортировано {sorted_count} файлов")

            # Если текущий файл не был найден — проверяем ещё раз
            if not found_in_current:
                found_folder_after_rescan = self.find_folder_by_newest_key(file_path, search_type)
                if found_folder_after_rescan:
                    print(f"✅ Текущий файл теперь сортируется в папку: '{found_folder_after_rescan}'")
                    return found_folder_after_rescan

        # Возвращаем результат для текущего файла
        if found_in_current:
            return found_in_current

        # Если ничего не найдено — спрашиваем, использовать ли папку
        print(f"\nКак поступить с текущим файлом '{filename}'?")
        action = input("  Создать папку для ключа (1) или выбрать другую папку (2)? ")
        if action == '1':
            return folder_name
        else:
            return None

    def find_folder_by_newest_key(self, file_path, added_search_type):
        """Вспомогательная функция для поиска по типу только что добавленного ключа"""
        filename = os.path.basename(file_path)
        file_ext = os.path.splitext(filename)[1].lower()

        if added_search_type == 'filename':
            return self.search_in_filename(filename)
        elif added_search_type == 'content':
            if file_ext in ['.xlsx', '.xls']:
                return self.search_exact_in_excel(file_path, filename)
            elif file_ext == '.pdf':
                return self.search_exact_in_pdf(file_path, filename)
            elif file_ext in ['.docx', '.doc']:
                return None
        return None


    def rescan_unsorted_by_search_type(self, new_search_key, search_type):
        """Ресортировка неотсортированных файлов ТОЛЬКО с новым ключом и указанным типом поиска"""
        print(f"\n🔄 Автоматическая ресортировка неотсортированных файлов с ключом: '{new_search_key}' (тип: {search_type})")
        sorted_count = 0
        unsorted_copy = self.unsorted_files.copy()

        for file_path, rel_path, organization in unsorted_copy:
            if not os.path.exists(file_path):
                self.unsorted_files.remove((file_path, rel_path, organization))
                continue

            filename = os.path.basename(file_path)
            file_ext = os.path.splitext(filename)[1].lower()

            found = False
            target_folder = None

            if search_type == 'filename':
                name_without_ext = os.path.splitext(filename)[0]
                clean_name = re.sub(r'[_\-.]', ' ', name_without_ext.lower())
                if new_search_key.lower() in clean_name:
                    target_folder = self.search_to_folder[new_search_key][0]
                    found = True
                    self.stats['name_matches'] += 1
            elif search_type == 'content':
                if file_ext in ['.xlsx', '.xls']:
                    wb = None
                    try:
                        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                        for sheet in wb.sheetnames:
                            ws = wb[sheet]
                            for row in ws.iter_rows(min_row=1, max_row=500, min_col=1, max_col=20, values_only=True):
                                row_texts = [str(cell).strip() for cell in row if cell]
                                if row_texts:
                                    row_line = ' '.join(row_texts)
                                    if new_search_key in row_line:
                                        target_folder = self.search_to_folder[new_search_key][0]
                                        found = True
                                        break
                                if found:
                                    break
                    except Exception as e:
                        self.log_detail(f"Ошибка Excel при ресортировке {filename}: {e}")
                    finally:
                        if wb:
                            wb.close()
                elif file_ext == '.pdf':
                    try:
                        import PyPDF2
                        with open(file_path, 'rb') as f:
                            pdf_reader = PyPDF2.PdfReader(f)
                            for page in pdf_reader.pages:
                                text = page.extract_text()
                                if text and new_search_key in text:
                                    target_folder = self.search_to_folder[new_search_key][0]
                                    found = True
                                    break
                    except Exception as e:
                        self.log_detail(f"Ошибка PDF при ресортировке {filename}: {e}")

            if found:
                print(f"   ✅ Найдено: {filename} → {target_folder}")
                source_date_part = self.extract_date_from_rel_path(rel_path)
                if self.move_file_to_folder(file_path, target_folder, organization, source_date_part):
                    self.stats['sorted'] += 1
                    self.stats['not_found'] -= 1
                    self.unsorted_files.remove((file_path, rel_path, organization))
                    sorted_count += 1
                else:
                    print(f"   ❌ Ошибка перемещения: {filename}")

        return sorted_count


    def get_file_preview(self, file_path, file_ext, max_chars=200):
        """Получение предпросмотра содержимого файла"""
        try:
            if file_ext in ['.xlsx', '.xls']:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheet = wb.active
                preview_lines = []
                for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                    row_data = [str(cell) for cell in row if cell]
                    if row: # <-- ИСПРАВЛЕНО: if row:
                        preview_lines.append(f"Строка {i}: {' | '.join(row_data[:5])}")
                wb.close()
                return '\n'.join(preview_lines)
            elif file_ext == '.pdf':
                try:
                    import PyPDF2
                    with open(file_path, 'rb') as f:
                        pdf_reader = PyPDF2.PdfReader(f)
                        text = pdf_reader.pages[0].extract_text()
                        return text[:max_chars] + ('...' if len(text) > max_chars else '')
                except Exception:
                    return "[Не удалось прочитать содержимое PDF]"
            else:
                return "[Просмотр содержимого недоступен для этого формата]"
        except Exception as e:
            return f"[Ошибка при чтении файла: {e}]"

    def preview_file_content(self, file_path, file_ext):
        """Предварительный просмотр содержимого файла"""
        try:
            print(f"\n📄 Просмотр содержимого файла:")
            print(f"   Путь: {file_path}")
            if file_ext in ['.xlsx', '.xls']:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheet = wb.active
                print(f"   Лист: {sheet.title}")
                print(f"   Размер: {sheet.max_row} строк, {sheet.max_column} колонок")
                print("\nПервые 10 строк:")
                for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                    row_data = [str(cell)[:50] for cell in row if cell]
                    if row: # <-- ИСПРАВЛЕНО: if row:
                        print(f"   {i:2}. {' | '.join(row_data)}")
                wb.close()
            elif file_ext == '.pdf':
                try:
                    import PyPDF2
                    with open(file_path, 'rb') as f:
                        pdf_reader = PyPDF2.PdfReader(f)
                        text = pdf_reader.pages[0].extract_text()
                        print(f"   Страниц: {len(pdf_reader.pages)}")
                        print("\nТекст первой страницы:")
                        lines = text.split('\n')
                        for i, line in enumerate(lines[:15], 1):
                            print(f"   {i:2}. {line[:80]}")
                except Exception:
                    print("   Не удалось прочитать содержимое PDF")
            else:
                print("   Просмотр содержимого недоступен для этого формата")
        except Exception as e:
            print(f"   Ошибка при просмотре файла: {e}")

    def scan_and_sort_by_filename(self):
        """Сканирование и сортировка неотсортированных файлов по имени файла"""
        print(f"\n🔄 Сканирование и сортировка по имени файла для {len(self.unsorted_files)} файлов...")
        sorted_count = 0
        unsorted_copy = self.unsorted_files.copy()

        for file_path, rel_path, organization in unsorted_copy:
            if not os.path.exists(file_path):
                self.unsorted_files.remove((file_path, rel_path, organization))
                continue

            filename = os.path.basename(file_path)
            folder_name = self.search_in_filename(filename)

            if folder_name:
                print(f"   ✅ Найдено совпадение по имени: {filename} → {folder_name}")
                source_date_part = self.extract_date_from_rel_path(rel_path)
                if self.move_file_to_folder(file_path, folder_name, organization, source_date_part):
                    self.stats['sorted'] += 1
                    self.stats['not_found'] -= 1
                    self.unsorted_files.remove((file_path, rel_path, organization))
                    sorted_count += 1
                else:
                    print(f"   ❌ Ошибка перемещения: {filename}")
            else:
                print(f"   ❌ Нет совпадений по имени: {filename}")

        print(f"✅ Отсортировано по имени файла: {sorted_count} файлов.")


    def rescan_unsorted_files(self):
        """Оригинальный метод — оставлен для совместимости (не используется в новой логике)"""
        print(f"\n🔄 РЕСОРТИРОВКА НЕОТСОРТИРОВАННЫХ ФАЙЛОВ")
        print(f"Неотсортированных файлов: {len(self.unsorted_files)}")
        sorted_count = 0
        unsorted_copy = self.unsorted_files.copy()

        for i, (file_path, rel_path, organization) in enumerate(unsorted_copy, 1):
            if not os.path.exists(file_path):
                self.unsorted_files.remove((file_path, rel_path, organization))
                continue

            folder_name = self.identify_report_type_with_filename(file_path)

            if folder_name:
                source_date_part = self.extract_date_from_rel_path(rel_path)
                if self.move_file_to_folder(file_path, folder_name, organization, source_date_part):
                    self.stats['sorted'] += 1
                    self.stats['not_found'] -= 1
                    self.unsorted_files.remove((file_path, rel_path, organization))
                    sorted_count += 1
            else:
                pass # Или логировать, если нужно

        return sorted_count

    def create_final_filename(self, original_filename, organization, destination_folder, source_date_part):
        """Создание окончательного имени файла по шаблону [Организация]_[Дата]_[Папка_назначения]"""
        # Очищаем название организации
        safe_org = re.sub(r'[<>:"/\\|?*]', '_', organization)
        safe_org = safe_org.strip('_')
        # Если организация неизвестна, используем плейсхолдер
        if safe_org == "Неизвестно" or not safe_org:
            safe_org = "Организация_Неизвестна"

        # Очищаем название папки назначения
        safe_dest_folder = re.sub(r'[<>:"/\\|?*]', '_', destination_folder)
        safe_dest_folder = safe_dest_folder.strip('_')
        if not safe_dest_folder:
             safe_dest_folder = "Папка_Назначения_Неизвестна"

        # Очищаем дату
        safe_date = re.sub(r'[<>:"/\\|?*]', '_', source_date_part)
        safe_date = safe_date.strip('_')
        if not safe_date:
            safe_date = "Дата_Неизвестна"

        # Получаем расширение файла
        _, ext = os.path.splitext(original_filename)

        # Формируем новое имя: [Организация]_[Дата]_[Папка_назначения]
        new_filename = f"{safe_org}_{safe_date}_{safe_dest_folder}{ext}"

        # Ограничиваем длину (Windows ограничение - 260 символов для полного пути, но ограничим имя файла)
        max_filename_length = 200
        if len(new_filename) > max_filename_length:
            # Обрезаем части, начиная с папки назначения, затем даты, затем организации
            excess = len(new_filename) - max_filename_length
            trunc_dest_len = max(len(safe_dest_folder) - excess, 5) # Минимум 5 символов
            safe_dest_folder = safe_dest_folder[:trunc_dest_len]

            new_filename = f"{safe_org}_{safe_date}_{safe_dest_folder}{ext}"
            if len(new_filename) > max_filename_length:
                excess = len(new_filename) - max_filename_length
                trunc_date_len = max(len(safe_date) - excess, 5)
                safe_date = safe_date[:trunc_date_len]

                new_filename = f"{safe_org}_{safe_date}_{safe_dest_folder}{ext}"
                if len(new_filename) > max_filename_length:
                    excess = len(new_filename) - max_filename_length
                    trunc_org_len = max(len(safe_org) - excess, 5)
                    safe_org = safe_org[:trunc_org_len]
                    new_filename = f"{safe_org}_{safe_date}_{safe_dest_folder}{ext}"

        return new_filename

    def extract_date_from_rel_path(self, rel_path):
        """Извлекает дату из относительного пути"""
        path_parts = rel_path.split(os.sep)
        if len(path_parts) >= 1:
            # Предполагаем, что дата находится в последней папке
            date_part = path_parts[-1]
            # Проверим, похожа ли на дату (например, YYYY-MM-DD_HHMM или YYYY-MM-DD)
            if re.match(r'^\d{4}-\d{2}-\d{2}(_\d+)?$', date_part):
                return date_part
        return "Дата_Не_Найдена"

    def move_file_to_folder(self, source_path, target_folder_name, organization, source_date_part=None):
        """ПЕРЕМЕЩЕНИЕ файла в целевую папку с новым именем по шаблону [Организация]_[Дата]_[Папка_назначения]"""
        # 🔥 Проверка существования файла перед перемещением
        if not os.path.exists(source_path):
            self.log_detail(f"  Файл не найден (уже перемещён?): {source_path}")
            return False

        safe_folder_name = re.sub(r'[<>:"/\\|?*]', '_', target_folder_name)
        safe_folder_name = safe_folder_name[:100].strip()
        target_dir = os.path.join(self.output_folder, safe_folder_name)
        os.makedirs(target_dir, exist_ok=True)
        self.found_folders.add(safe_folder_name)

        original_filename = os.path.basename(source_path)
        # Извлекаем дату из пути источника, если не передана
        if source_date_part is None:
            # Путь может быть: Организации_и_письма/Название_организации/2024-01-15_1430/файл.xlsx
            # Извлекаем вторую часть от конца (дату_время)
            rel_path_from_output = os.path.relpath(source_path, start=self.source_folder)
            path_parts = rel_path_from_output.split(os.sep)
            if len(path_parts) >= 2:
                source_date_part = path_parts[-2] # Вторая часть от конца
            else:
                source_date_part = "Дата_Не_Найдена"

        # Создаем новое имя файла по шаблону
        final_filename = self.create_final_filename(original_filename, organization, target_folder_name, source_date_part)
        target_path = os.path.join(target_dir, final_filename)

        # Если файл уже существует, добавляем номер
        counter = 1
        base_name, ext = os.path.splitext(target_path)
        while os.path.exists(target_path):
            target_path = f"{base_name}_{counter}{ext}"
            counter += 1

        try:
            shutil.move(source_path, target_path)
            self.stats['moved'] += 1
            log_msg = f"  ПЕРЕМЕЩЕН в: {safe_folder_name}/{os.path.basename(target_path)}"
            if counter > 1:
                log_msg += f" (переименован с {original_filename})"
            print(log_msg)
            self.log_detail(log_msg)
            return True
        except Exception as e:
            error_msg = f"  ❌ Ошибка перемещения {original_filename}: {e}"
            print(error_msg)
            self.log_detail(f"  Ошибка перемещения {original_filename}: {e}")
            self.stats['errors'] += 1
            return False

    def scan_all_files(self):
        """Сканирование всех файлов"""
        print(f"\n🔍 Сканирование папки: {self.source_folder}")
        all_files = []
        for root, dirs, files in os.walk(self.source_folder):
            for file in files:
                file_ext = os.path.splitext(file)[1].lower()
                if file_ext in self.supported_formats:
                    file_path = os.path.join(root, file)
                    rel_path = os.path.relpath(root, self.source_folder)
                    all_files.append((file_path, rel_path))

        self.stats['total_files'] = len(all_files)
        print(f"✅ Найдено файлов: {self.stats['total_files']}")
        self.all_files_original = all_files.copy()
        return all_files

    def process_file(self, file_info):
        """Обработка одного файла"""
        file_path, rel_path = file_info
        try:
            self.stats['processed'] += 1
            current_num = self.stats['processed']
            total_files = self.stats['total_files']

            if current_num % 50 == 0:
                print(f"📊 [{current_num:4}/{total_files:4}] "
                      f"Отсортировано: {self.stats['sorted']:4} | "
                      f"Точных совпадений: {self.stats['exact_matches']:4} | "
                      f"По имени: {self.stats['name_matches']:4} | "
                      f"Не найдено: {self.stats['not_found']:4}")

            filename = os.path.basename(file_path)

            organization = self.extract_organization_from_path(file_path, rel_path)

            folder_name = self.identify_report_type(file_path)

            if folder_name:
                self.stats['exact_matches'] += 1
                # Передаем rel_path для извлечения даты
                if self.move_file_to_folder(file_path, folder_name, organization, self.extract_date_from_rel_path(rel_path)):
                    self.stats['sorted'] += 1
                    return (file_path, folder_name, True, "Успешно перемещен", organization)
                else:
                    return (file_path, None, False, "Ошибка перемещения", organization)
            else:
                if self.interactive:
                    self.unsorted_files.append((file_path, rel_path, organization))
                    self.stats['not_found'] += 1
                    return (file_path, None, False, "Ожидает интерактивной обработки", organization)
                else:
                    self.stats['not_found'] += 1
                    # Передаем rel_path для извлечения даты
                    if self.move_file_to_folder(file_path, "НЕ_СОРТИРОВАННЫЕ", organization, self.extract_date_from_rel_path(rel_path)):
                        return (file_path, "НЕ_СОРТИРОВАННЫЕ", True, "Перемещен в НЕ_СОРТИРОВАННЫЕ", organization)
                    else:
                        return (file_path, None, False, "Ошибка перемещения в НЕ_СОРТИРОВАННЫЕ", organization)
        except Exception as e:
            self.stats['errors'] += 1
            error_msg = f"Критическая ошибка обработки {file_path}: {e}"
            print(f"❌ {error_msg}")
            self.log_detail(error_msg)
            return (file_path, None, False, str(e), "Неизвестно")

    def process_interactive_files(self):
        """Обработка файлов в интерактивном режиме"""
        print(f"\n🔧 ИНТЕРАКТИВНЫЙ РЕЖИМ")
        print(f"Файлов для ручной обработки: {len(self.unsorted_files)}")
        print("="*60)
        unsorted_copy = self.unsorted_files.copy()

        for i, (file_path, rel_path, organization) in enumerate(unsorted_copy, 1):
            # Проверяем, не был ли файл уже перемещен
            if not os.path.exists(file_path):
                print(f"\n⚠️  Файл уже перемещен, пропускаем")
                continue

            filename = os.path.basename(file_path)
            file_ext = os.path.splitext(filename)[1].lower()

            print(f"\n📋 Файл {i}/{len(unsorted_copy)}: {filename}")
            print(f"   Организация: {organization}")

            folder_choice = self.get_interactive_choice(filename, file_ext, file_path, organization)

            if folder_choice:
                self.stats['interactive_choices'] += 1
                # Извлекаем дату из rel_path
                source_date_part = self.extract_date_from_rel_path(rel_path)
                if self.move_file_to_folder(file_path, folder_choice, organization, source_date_part):
                    self.stats['sorted'] += 1
                    self.stats['not_found'] -= 1
                    if (file_path, rel_path, organization) in self.unsorted_files:
                        self.unsorted_files.remove((file_path, rel_path, organization))
                else:
                    # Если ошибка перемещения, оставляем в исходной папке
                    print(f"  ⚠️  Файл оставлен в исходной папке: {file_path}")
            else:
                print(f"  ⚠️  Файл пропущен: {filename}")
                # Файл остается в исходной папке и в списке неотсортированных

    def cleanup_empty_txt_dirs(self):
        """Удаление папок в корне исходной директории, если в них только один .txt файл"""
        print(f"\n🔍 Проверка папок в исходной директории '{self.source_folder}' на наличие только одного .txt файла...")
        try:
            items_in_source = os.listdir(self.source_folder)
        except OSError as e:
            print(f"❌ Ошибка доступа к исходной директории '{self.source_folder}': {e}")
            self.log_detail(f"Ошибка доступа к исходной директории '{self.source_folder}': {e}")
            return

        deleted_dirs_count = 0
        for item in items_in_source:
            item_path = os.path.join(self.source_folder, item)
            # Проверяем, что это директоря
            if os.path.isdir(item_path):
                try:
                    dir_contents = os.listdir(item_path)
                    # Считаем .txt файлы и другие элементы
                    txt_files = [f for f in dir_contents if os.path.isfile(os.path.join(item_path, f)) and f.lower().endswith('.txt')]
                    other_items = [f for f in dir_contents if f not in txt_files]

                    # Если есть только один .txt файл и нет других файлов/папок
                    if len(txt_files) == 1 and len(other_items) == 0:
                        print(f"\n🗑️  Найдена папка '{item}' с единственным .txt файлом '{txt_files[0]}'. Удаляем папку.")
                        shutil.rmtree(item_path)
                        print(f"✅ Папка '{item}' удалена.")
                        self.log_detail(f"Папка '{item}' в исходной директории была удалена, так как содержала только один .txt файл: {txt_files[0]}")
                        deleted_dirs_count += 1
                    else:
                        # Только для отладки, можно убрать
                        # print(f"   Папка '{item}' содержит {len(txt_files)} .txt файлов и {len(other_items)} других элементов. Не удаляется.")
                        pass
                except OSError as e:
                    print(f"❌ Ошибка при проверке папки '{item_path}': {e}")
                    self.log_detail(f"Ошибка при проверке папки '{item_path}': {e}")

        if deleted_dirs_count == 0:
            print(f"📋 Ни одной папки в '{self.source_folder}' не было удалено по условию.")
        else:
            print(f"✅ Удалено {deleted_dirs_count} папок по условию.")


    def process_all_files(self, max_workers=4):
        """Обработка всех файлов"""
        if not self.load_report_names():
            return False

        all_files = self.scan_all_files()
        if not all_files:
            print("⚠️ Файлы не найдены!")
            return False

        print(f"\n🚀 Начинаем обработку {len(all_files)} файлов...")
        print("="*60)
        print("⚠️  ВНИМАНИЕ: Ищем ТОЛЬКО в содержимом файлов (при первичной обработке)")
        print("⚠️  Имена файлов игнорируются на первом этапе!")
        print("⚠️  К именам файлов будет добавлен отправитель")
        print("⚠️  Файлы ПЕРЕМЕЩАЮТСЯ (не копируются)!")
        print("="*60)

        # Создаем папку для неотсортированных
        unsorted_folder = os.path.join(self.output_folder, "НЕ_СОРТИРОВАННЫЕ")
        os.makedirs(unsorted_folder, exist_ok=True)
        self.found_folders.add("НЕ_СОРТИРОВАННЫЕ")

        # Обработка файлов
        results = []

        if self.interactive:
            # В интерактивном режиме используем только один поток
            print("\n🔄 Обработка файлов в однопоточном режиме (интерактивный режим)...")
            for file_info in all_files:
                file_path, rel_path = file_info
                filename = os.path.basename(file_path)
                # Обновляем прогресс
                self.stats['processed'] += 1
                if self.stats['processed'] % 10 == 0:
                    print(f"📊 [{self.stats['processed']:4}/{len(all_files):4}] "
                          f"Отсортировано: {self.stats['sorted']:4} "
                          f"Неотсортировано: {len(self.unsorted_files):4}")

                organization = self.extract_organization_from_path(file_path, rel_path)
                source_date_part = self.extract_date_from_rel_path(rel_path)

                # Сначала пытаемся автоматически определить ТОЛЬКО по содержимому
                folder_name = self.identify_report_type(file_path)

                if folder_name:
                    # Автоматическое перемещение
                    self.stats['exact_matches'] += 1
                    if self.move_file_to_folder(file_path, folder_name, organization, source_date_part):
                        self.stats['sorted'] += 1
                        results.append((file_path, folder_name, True, "Успешно перемещен", organization))
                    else:
                        results.append((file_path, None, False, "Ошибка перемещения", organization))
                else:
                    # Добавляем в список для интерактивной обработки
                    self.unsorted_files.append((file_path, rel_path, organization))
                    results.append((file_path, None, False, "Ожидает интерактивной обработки", organization))
                    self.stats['not_found'] += 1

            # Обработка файлов в интерактивном режиме
            if self.unsorted_files:
                self.process_interactive_files()
        else:
            # Неинтерактивный режим - используем многопоточность
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_file = {executor.submit(self.process_file, file_info): file_info
                                  for file_info in all_files}
                for future in as_completed(future_to_file):
                    try:
                        result = future.result()
                        results.append(result)
                    except Exception as e:
                        error_msg = f"Ошибка в потоке: {e}"
                        print(f"❌ {error_msg}")
                        self.log_detail(error_msg)

        # --- НОВОЕ ---
        # Выполняем очистку после завершения основной сортировки
        self.cleanup_empty_txt_dirs()

        # Генерация отчета
        self.generate_report(results)
        return True

    def generate_report(self, results):
        """Генерация итогового отчета"""
        report_file = os.path.join(self.output_folder, "ИТОГОВЫЙ_ОТЧЕТ.txt")

        # Группируем результаты
        report_stats = {}
        organizations_used = set()
        for file_path, folder_name, success, message, organization in results:
            if success and folder_name:
                report_stats[folder_name] = report_stats.get(folder_name, 0) + 1
            if organization and organization != "Неизвестно":
                organizations_used.add(organization)

        with open(report_file, 'w', encoding='utf-8') as f:
            f.write("="*80 + "\n")
            f.write("ИТОГОВЫЙ ОТЧЕТ СОРТИРОВКИ\n")
            f.write("="*80 + "\n")
            f.write(f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}\n")
            f.write(f"Исходная папка: {self.source_folder}\n")
            f.write(f"Выходная папка: {self.output_folder}\n")
            f.write(f"Файл с настройками: {self.report_names_file}\n")
            f.write(f"Интерактивный режим: {'Да' if self.interactive else 'Нет'}\n")
            if self.interactive:
                f.write(f"Новых ключей добавлено: {self.stats['new_keys_added']}\n")
            f.write("⚠️  РЕЖИМ ПОИСКА (при первичной обработке): ТОЛЬКО В СОДЕРЖИМОМ ФАЙЛОВ\n")
            f.write("⚠️  ИМЕНА ФАЙЛОВ ИГНОРИРУЮТСЯ!\n")
            f.write("⚠️  К именам файлов добавлен отправитель (если известен)\n")
            f.write("⚠️  ФАЙЛЫ ПЕРЕМЕЩАЮТСЯ, А НЕ КОПИРУЮТСЯ!\n")
            f.write("="*80 + "\n")
            f.write("СТАТИСТИКА\n")
            f.write("="*80 + "\n")
            f.write(f"Всего файлов: {self.stats['total_files']}\n")
            f.write(f"Обработано: {self.stats['processed']}\n")
            f.write(f"Успешно перемещено: {self.stats['moved']}\n")
            f.write(f"Точных совпадений в содержимом: {self.stats['exact_matches']}\n")
            f.write(f"Совпадений по имени файла (после добавления ключей): {self.stats['name_matches']}\n")
            if self.interactive:
                f.write(f"Интерактивных выборов: {self.stats['interactive_choices']}\n")
                f.write(f"Добавлено новых ключей: {self.stats['new_keys_added']}\n")
            f.write(f"Не распознано: {self.stats['not_found']}\n")
            f.write(f"Ошибок: {self.stats['errors']}\n")
            if self.interactive and self.unsorted_files:
                f.write(f"⚠️  Осталось неотсортированных файлов: {len(self.unsorted_files)}\n")

            # Детализация по папкам
            if report_stats:
                f.write("="*80 + "\n")
                f.write("РАСПРЕДЕЛЕНИЕ ФАЙЛОВ ПО ПАПКАМ\n")
                f.write("="*80 + "\n")
                # Сортируем по количеству файлов
                sorted_stats = sorted(report_stats.items(), key=lambda x: x[1], reverse=True)
                for folder_name, count in sorted_stats:
                    f.write(f"📁 {folder_name}: {count} файл(ов)\n")

            # Информация об организациях
            if organizations_used:
                f.write("\n" + "="*80 + "\n")
                f.write("ИСПОЛЬЗОВАННЫЕ ОРГАНИЗАЦИИ-ОТПРАВИТЕЛИ\n")
                f.write("="*80 + "\n")
                for org in sorted(organizations_used):
                    f.write(f"🏢 {org}\n")

            # Файлы, оставшиеся в исходной папке
            remaining_files = [(file_path, message) for file_path, folder_name, success, message, organization
                               in results if not success or not folder_name or message == "Ожидает интерактивной обработки"]
            if remaining_files:
                f.write("\n" + "="*80 + "\n")
                f.write("ФАЙЛЫ, ОСТАВШИЕСЯ В ИСХОДНОЙ ПАПКЕ\n")
                f.write("="*80 + "\n")
                for file_path, message in remaining_files[:50]:  # Ограничиваем вывод
                    filename = os.path.basename(file_path)
                    f.write(f"❌ {filename}: {message}\n")
                if len(remaining_files) > 50:
                    f.write(f"\n... и еще {len(remaining_files) - 50} файлов\n")

            f.write("\n" + "="*80 + "\n")
            f.write("ВНИМАНИЕ\n")
            f.write("="*80 + "\n")
            f.write("1. Файлы были ПЕРЕМЕЩЕНЫ из исходной папки\n")
            f.write("2. Исходные файлы больше не существуют в исходном расположении\n")
            f.write("3. Для отмены операции потребуется восстановление из бэкапа\n")
            f.write("4. Всегда делайте бэкап перед запуском сортировки!\n")

            if self.interactive and self.stats['new_keys_added'] > 0:
                f.write(f"✅ Добавлено {self.stats['new_keys_added']} новых ключей поиска\n")
                f.write("   Ключи сохранены в файле настроек\n")
                f.write("   Их можно использовать при следующем запуске\n")

            f.write("✅ Сортировка завершена!\n")

        print(f"\n📊 Отчет сохранен: {report_file}")
        print(f"📝 Подробный лог: {self.log_file}")

        # Вывод краткой статистики в консоль
        print("\n" + "="*60)
        print("ИТОГИ:")
        print(f"📁 Всего файлов: {self.stats['total_files']}")
        print(f"✅ Перемещено: {self.stats['moved']}")
        print(f"🎯 Точных совпадений в содержимом: {self.stats['exact_matches']}")
        print(f"🎯 Совпадений по имени файла (после добавления ключей): {self.stats['name_matches']}")
        if self.interactive:
            print(f"👤 Интерактивных выборов: {self.stats['interactive_choices']}")
            print(f"➕ Новых ключей добавлено: {self.stats['new_keys_added']}")
            print(f"❓ Осталось неотсортированных: {len(self.unsorted_files)}")
        else:
            print(f"❓ Не распознано/оставлено: {self.stats['not_found']}")
        print(f"⚠️  Ошибок: {self.stats['errors']}")
        print("="*60)

def main():
    parser = argparse.ArgumentParser(description='Сортировка отчетов по содержимому файлов')
    parser.add_argument('--source', required=True, help='Исходная папка с файлами')
    parser.add_argument('--output', required=True, help='Выходная папка для сортировки')
    parser.add_argument('--config', required=True, help='Файл с названиями отчетов и ключами поиска')
    parser.add_argument('--interactive', action='store_true', help='Интерактивный режим')
    parser.add_argument('--workers', type=int, default=4, help='Количество потоков (по умолчанию: 4)')

    args = parser.parse_args()

    print("="*80)
    print("📁 СОРТИРОВЩИК ОТЧЕТОВ ПО СОДЕРЖИМОМУ ФАЙЛОВ")
    print("="*80)
    print(f"Исходная папка: {args.source}")
    print(f"Выходная папка: {args.output}")
    print(f"Файл настроек: {args.config}")
    print(f"Интерактивный режим: {'Да' if args.interactive else 'Нет'}")
    print(f"Потоков обработки: {args.workers}")
    print("="*80)
    print("⚠️  ВНИМАНИЕ: Файлы будут ПЕРЕМЕЩЕНЫ, а не скопированы!")
    print("⚠️  Рекомендуется сделать резервную копию перед запуском!")
    print("="*80)

    if args.interactive:
        print("\n🔄 ИНТЕРАКТИВНЫЙ РЕЖИМ ВКЛЮЧЕН")
        print("Для каждого нераспознанного файла будет запрошено действие.")
        print("Можно добавлять новые ключи поиска (в содержимом или в имени файла) и выполнять ресортировку.")
        confirm = input("\nПродолжить? (да/НЕТ): ").strip().lower()
        if confirm != 'да':
            print("Отменено пользователем.")
            return

    if not os.path.exists(args.source):
        print(f"❌ Исходная папка не существует: {args.source}")
        return

    if not os.path.exists(args.config):
        print(f"❌ Файл настроек не существует: {args.config}")
        return

    sorter = ReportSorter(
        source_folder=args.source,
        output_folder=args.output,
        report_names_file=args.config,
        interactive=args.interactive
    )

    try:
        success = sorter.process_all_files(max_workers=args.workers if not args.interactive else 1)
        if success:
            print("\n✅ Сортировка завершена успешно!")
            print(f"\n📁 Результаты в папке: {args.output}")
            # Показываем созданные папки
            if os.path.exists(args.output):
                folders = [d for d in os.listdir(args.output)
                           if os.path.isdir(os.path.join(args.output, d))]
                if folders:
                    print(f"\n📂 Создано папок: {len(folders)}")
                    print("Основные папки:")
                    for folder in sorted(folders)[:10]:  # Показываем первые 10
                        print(f"  📁 {folder}")
                    if len(folders) > 10:
                        print(f"  ... и еще {len(folders) - 10} папок")
        else:
            print("\n❌ Сортировка завершена с ошибками!")

    except KeyboardInterrupt:
        print("\n⚠️  Процесс прерван пользователем!")
        print("⚠️  Частичные результаты сохранены.")
    except Exception as e:
        print(f"\n❌ Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()